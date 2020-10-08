import openpyxl
import settings
import datetime


def send_xlsx(a):
    wb = openpyxl.Workbook()

    wb.create_sheet(title='Отчет', index=0)

    sheet = wb['Отчет']
    now = datetime.datetime.now()
    data = now.strftime('%H-%M-%S_%d.%m.%Y')
    len_keys = 1
    for i in a:
        if len(i) > len_keys:
            len_keys = len(i)

    sheet.column_dimensions['A'].width = len_keys + 2
    for (row, i) in zip(range(1, len(a) + 1), a):
        fio = i
        summ = a[i]
        sheet.cell(column=1, row=row, value=fio)
        sheet.cell(column=2, row=row, value=summ)
    try:
        file = settings.xlsx_path + 'report_' + str(data) + '.xlsx'
        wb.save(file)
    except PermissionError:
        file = None
        pass
    return file
